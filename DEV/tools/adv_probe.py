"""One-shot schema probe for the SEC Form ADV ("Information about Registered
Investment Advisers") data set. Runs on a GitHub runner (which reaches sec.gov).

It downloads the latest monthly ADV file, lists what's inside, and prints the
column headers + a sample row so we can build the enricher against the REAL
schema (which fields carry AUM, private-fund counts, administrator/auditor,
registration status) rather than guessing.

Writes briefs/adv_schema.txt and prints to stdout. Always exits 0.
"""

from __future__ import annotations

import io
import zipfile
from pathlib import Path

import urllib.request

UA = "Coremont Clarion Prospecting malex.clark@gmail.com"

# Candidate URLs (most recent monthly first); from the data.gov catalog listing.
CANDIDATE_URLS = [
    "https://www.sec.gov/files/investment/data/other/information-about-registered-investment-advisers-exempt-reporting-advisers/ia060126.zip",
    "https://www.sec.gov/files/investment/data/information-about-registered-investment-advisers-exempt-reporting-advisers/ia060126.zip",
    "https://www.sec.gov/files/investment/data/other/information-about-registered-investment-advisers-exempt-reporting-advisers/ia050126.zip",
]

KEYWORDS = (
    "crd", "sec", "name", "aum", "asset", "regulatory", "private fund", "fund",
    "administrator", "auditor", "custodian", "prime", "gross", "employees",
    "registration", "status", "date", "7.b", "7b", "schedule", "gav",
)


def _get(url: str) -> bytes | None:
    try:
        req = urllib.request.Request(url, headers={"User-Agent": UA})
        with urllib.request.urlopen(req, timeout=120) as r:
            return r.read()
    except Exception as e:  # noqa: BLE001
        print(f"  download failed: {url} -> {e}")
        return None


def _headers_of(name: str, data: bytes) -> list[str]:
    low = name.lower()
    if low.endswith(".csv") or low.endswith(".txt"):
        text = data[:200_000].decode("latin-1", errors="replace")
        first = text.splitlines()[0] if text else ""
        delim = "," if first.count(",") >= first.count("\t") else "\t"
        return [h.strip().strip('"') for h in first.split(delim)]
    if low.endswith(".xlsx"):
        try:
            import openpyxl  # type: ignore

            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
            ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                return [str(c) for c in row if c is not None]
        except Exception as e:  # noqa: BLE001
            return [f"<xlsx read error: {e}>"]
    return []


def main() -> None:
    out: list[str] = ["SEC Form ADV data-set schema probe", "=" * 60]
    blob = None
    used = None
    for url in CANDIDATE_URLS:
        out.append(f"trying: {url}")
        blob = _get(url)
        if blob:
            used = url
            out.append(f"  OK {len(blob):,} bytes")
            break
    if not blob:
        out.append("No ADV file could be downloaded.")
        _emit(out)
        return

    out.append(f"\nUsing: {used}\n")
    try:
        zf = zipfile.ZipFile(io.BytesIO(blob))
    except zipfile.BadZipFile:
        out.append("Downloaded file is not a zip; first 200 bytes:")
        out.append(repr(blob[:200]))
        _emit(out)
        return

    for info in zf.infolist():
        out.append(f"--- {info.filename}  ({info.file_size:,} bytes) ---")
        try:
            data = zf.read(info.filename)
        except Exception as e:  # noqa: BLE001
            out.append(f"  read error: {e}")
            continue
        cols = _headers_of(info.filename, data)
        out.append(f"  {len(cols)} columns")
        # Show columns that look relevant to our signals.
        relevant = [c for c in cols if any(k in c.lower() for k in KEYWORDS)]
        out.append("  RELEVANT columns:")
        for c in relevant[:60]:
            out.append(f"    - {c}")
        out.append(f"  (all {len(cols)} headers): {cols[:40]}")
    _emit(out)


def _emit(lines: list[str]) -> None:
    text = "\n".join(lines)
    print(text)
    try:
        dest = Path(__file__).resolve().parents[1] / "briefs" / "adv_schema.txt"
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_text(text, encoding="utf-8")
        print(f"\nwrote {dest}")
    except OSError as e:
        print(f"(could not write file: {e})")


if __name__ == "__main__":
    main()
